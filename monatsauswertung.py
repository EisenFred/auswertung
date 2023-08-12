import pandas as pd
from openpyxl.styles import PatternFill
from load_csv import load_csv
import datetime
import openpyxl

# Spalte vor das Datum mit dem Wochentag

EXCEL_FILENAME = 'output.xlsx'
CSV_FILENAME = "SB56_t30_Ferien.csv"
MONTH = 7
YEAR = 2023

def generate_dates(year, month):
    num_days = (datetime.date(year, month + 1, 1) - datetime.date(year, month, 1)).days
    dates = [datetime.date(year, month, day).strftime('%d-%m-%Y') for day in range(1, num_days + 1)]
    return dates

def monatsauswertung():
    data = load_csv(CSV_FILENAME)
    columns = ['Linie', 'Soll', 'Ist', 'Verspaetung', 'Datum', 'Wochentag']
    df_csv = pd.DataFrame(data, columns=columns)
    df_csv['Soll'] = pd.to_datetime(df_csv['Soll'], format='%H:%M')
    df_csv['Ist'] = pd.to_datetime(df_csv['Ist'], format='%H:%M')
    df_csv['Datum'] = pd.to_datetime(df_csv['Datum'], format='%d.%m.%Y', dayfirst=True)
    start_date = pd.to_datetime(f'2023-{MONTH:02d}-01')
    end_date = pd.to_datetime(f'2023-{MONTH:02d}-31')
    df_csv = df_csv[(df_csv['Datum'] >= start_date) & (df_csv['Datum'] <= end_date)]

    rows = 32
    columns2 = ['07:16','07:46','08:16','08:46','09:16','09:46','10:16','10:46','11:16','11:46','12:16','12:46','13:16','13:46','14:16','14:46','15:16','15:46','16:16','16:46','17:16','17:46','18:16','18:46','19:16','19:46','20:16','20:46','21:16','21:46','22:16','22:46',]
    df_auswertung = pd.DataFrame(index=range(rows), columns=columns2)

    for index_day_in_month in range(1,32):
        date_day = pd.to_datetime(f'2023-{MONTH:02d}-{index_day_in_month:02d}')
        df_thisDay = df_csv[(df_csv['Datum'] == date_day)]
        df_thisDay = df_thisDay.copy() # Verhindert eine Warnung die durch die beiden folgenden Zeilen sonst entstehen würden.
        df_thisDay['Ist'] = df_thisDay['Ist'].dt.strftime('%H:%M')
        df_thisDay['Soll'] = df_thisDay['Soll'].dt.strftime('%H:%M')
        df_thisDay['Verspaetung'] = df_thisDay['Verspaetung'].astype(int)
        df_thisDay = df_thisDay.sort_values(by='Soll')
        df_thisDay = df_thisDay.reset_index(drop=True)
        index_df_thisDay = 0
        if not df_thisDay.empty:
            for time_day in range(0,32):
                if not index_df_thisDay >= len(df_thisDay):
                    if df_auswertung.columns[time_day] == df_thisDay.loc[index_df_thisDay,'Soll']:
                        df_auswertung.iloc[index_day_in_month-1,time_day] = df_thisDay.loc[index_df_thisDay, 'Verspaetung']
                        index_df_thisDay += 1
    array = generate_dates(YEAR, MONTH)
    new_column = pd.Series(array, name='Datum')
    df_auswertung.insert(0, 'Datum', new_column)

    df_auswertung.to_excel(EXCEL_FILENAME, index=False)

    workbook = openpyxl.load_workbook(EXCEL_FILENAME)
    worksheet = workbook['Sheet1']
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 11
    for index in range(2,34):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 5

    data_range = 'B2:AG32'

    for row in worksheet[data_range]:
        for cell in row:
            if cell.value == 0:
                cell.fill = PatternFill(start_color='9bbb59', end_color='9bbb59', fill_type='solid')
            if cell.value == 1:
                cell.fill = PatternFill(start_color='b6be3c', end_color='b6be3c', fill_type='solid')
            if cell.value == 2:
                cell.fill = PatternFill(start_color='d2c21e', end_color='d2c21e', fill_type='solid')
            if cell.value == 3:
                cell.fill = PatternFill(start_color='eec600', end_color='eec600', fill_type='solid')
            if cell.value == 4:
                cell.fill = PatternFill(start_color='d78b26', end_color='d78b26', fill_type='solid')
            if cell.value is not None and cell.value >= 5:
                cell.fill = PatternFill(start_color='c0504d', end_color='c0504d', fill_type='solid')
            if cell.value is not None and cell.value >= 10:
                cell.fill = PatternFill(start_color='8e1e1b', end_color='8e1e1b', fill_type='solid')
    workbook.save(EXCEL_FILENAME)

monatsauswertung()





